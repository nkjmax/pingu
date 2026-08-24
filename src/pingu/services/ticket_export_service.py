"""
Keeps a local .xlsx mirror of the tickets table -- NOT the source of
truth (SQLite is that, as always), just a convenient exportable view for
non-technical review. Updated incrementally as tickets change state.
Never pushed to git (see .gitignore).

openpyxl is a SYNCHRONOUS library -- calling it directly from an async
handler would block the entire event loop for however long the file
read/write takes, defeating the whole point of everything else in this
bot being non-blocking. Every openpyxl call here goes through
asyncio.to_thread() so it runs on a background thread instead.

Worth knowing as this scales: each write does a full load + full save of
the whole workbook (openpyxl doesn't support partial/incremental writes),
so this gets slower as the ticket count grows -- fine for the hundreds-
of-tickets range, would need a different approach (e.g. batched writes,
or a real spreadsheet-less export) if ticket volume ever got into the
tens of thousands.
"""

import asyncio
import os
from datetime import datetime

from openpyxl import Workbook, load_workbook

EXCEL_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "tickets_export.xlsx"
)

COLUMNS = [
    "Ticket Number", "Category", "Subcategory", "Type", "Submitter (Discord ID)",
    "Description", "Status", "Created At (SGT)", "Closed At (SGT)", "Closed By (Discord ID)",
]


def _fmt_ts(unix_ts):
    if not unix_ts:
        return ""
    from dateutil.tz import gettz
    return datetime.fromtimestamp(unix_ts, tz=gettz("Asia/Singapore")).strftime("%Y-%m-%d %H:%M:%S")


def _upsert_row_sync(ticket: dict):
    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tickets"
        ws.append(COLUMNS)
    else:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Tickets"] if "Tickets" in wb.sheetnames else wb.active

    row_data = [
        ticket["ticket_number"],
        ticket["category"],
        ticket["subcategory"] or "",
        ticket["ticket_type"],
        str(ticket["user_id"]),
        ticket["body"],
        ticket["status"],
        _fmt_ts(ticket["created_at"]),
        _fmt_ts(ticket["closed_at"]),
        str(ticket["closed_by"]) if ticket["closed_by"] else "",
    ]

    existing_row = None
    for row in ws.iter_rows(min_row=2, max_col=1):
        if row[0].value == ticket["ticket_number"]:
            existing_row = row[0].row
            break

    if existing_row:
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=existing_row, column=col, value=value)
    else:
        ws.append(row_data)

    wb.save(EXCEL_PATH)


async def upsert_ticket_row(ticket) -> None:
    """ticket is a db row (or anything dict-like) with the tickets table's columns."""
    ticket_dict = dict(ticket)
    await asyncio.to_thread(_upsert_row_sync, ticket_dict)